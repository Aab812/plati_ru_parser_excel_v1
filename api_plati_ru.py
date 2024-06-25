import os
import subprocess
import tkingister as tk
from tkinter import ttk
from datetime import datetime
import openpyxl
import aiohttp
import asyncio
from openpyxl.styles import PatternFill
import json

async def fetch_data(session, query, page_size, page_end):
    url = f"https://plati.io/api/search.ashx?query={query}&pagesize={page_size}&pagenum={page_end}&visibleOnly=true&response=json"
    headers = {'Content-Type': 'application/json'}
    async with session.get(url, headers=headers) as response:
        data_text = await response.text()
        try:
            data = json.loads(data_text)
        except json.JSONDecodeError:
            print("Ошибка при декодировании JSON")
            data = {}
        return data.get('items', [])

def extract_account_info(description, title):
    keywords = ["аккаунт", "account", "аренда", "offline", "Оффлайн", "Онлайн", "п2", "п3", " п2 ", "п2-п3"]
    account_info = "Упоминание аккаунта" if any(keyword in title.lower() for keyword in keywords) else ""
    return account_info

async def add_data_to_excel(session, params):
    query = params['query']
    page_size = params['page_size']
    page_end = params['page_end']
    min_price = params.get('min_price')
    max_price = params.get('max_price')

    current_time = datetime.now()
    month_folder = current_time.strftime("%B")
    os.makedirs(month_folder, exist_ok=True)
    folder_path = os.path.abspath(month_folder)
    file_name = current_time.strftime(f"{month_folder}/%d_%m_%Y_%H_%M_{query}.xlsx")

    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
        all_items = await fetch_all_items(session, query, page_size, page_end)
        filtered_items = filter_items(all_items, min_price, max_price)
        sorted_items = sorted(filtered_items, key=lambda x: x.get("price_rur", 0), reverse=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "Название", "Цена", "Ссылка", "Платформа", "Продавец", "Рейтинг продавца", "Положительные отзывы",
                   "Отрицательные отзывы", "Возвраты", "Упоминание аккаунта"])

        for idx, item in enumerate(sorted_items, start=1):
            name = item.get("name", "")
            price = item.get("price_rur", "")
            url = item.get("url", "")
            title = name.lower()
            platform = extract_platform(title)
            account_info = extract_account_info(item.get("description", ""), title)
            ws.append(
                [item.get("seller_id", ""), name, price, url, platform, item.get("seller_name", ""),
                 item.get("seller_rating", ""), item.get("count_positiveresponses", ""),
                 item.get("count_negativeresponses", ""), item.get("count_returns", ""), account_info])

            fill = None
            seller_rating = item.get("seller_rating", "")
            if seller_rating < 100:
                fill = PatternFill(start_color="fdc0c8", end_color="fdc0c8", fill_type="solid")
            elif 100 <= seller_rating < 500:
                fill = PatternFill(start_color="fde890", end_color="fde890", fill_type="solid")
            else:
                fill = PatternFill(start_color="c0ecc7", end_color="c0ecc7", fill_type="solid")

            for cell in ws[idx + 1]:
                cell.fill = fill

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        ws.auto_filter.ref = ws.dimensions
        wb.save(file_name)

    return folder_path

async def fetch_all_items(session, query, page_size, page_end):
    tasks = [fetch_data(session, query, page_size, page_num) for page_num in range(1, page_end + 1)]
    return [item for sublist in await asyncio.gather(*tasks) for item in sublist]

def filter_items(items, min_price, max_price):
    if min_price is not None:
        items = [item for item in items if item.get("price_rur", 0) >= min_price]

    if max_price is not None:
        items = [item for item in items if item.get("price_rur", 0) <= max_price]

    return items

def extract_platform(title):
    platforms = ["Xbox", "PC", "PS", "Steam"]
    return next((platform for platform in platforms if platform.lower() in title), "")

def open_excel_folder():
    query = query_combobox.get()
    min_price = int(min_price_entry.get() or 0)
    max_price = int(max_price_entry.get() or 0)
    params = {'query': query, 'page_size': 10, 'page_end': 10, 'min_price': min_price, 'max_price': max_price}
    asyncio.run(add_data_to_excel(None, params))

def run_query():
    query = query_combobox.get()
    min_price = min_price_entry.get()
    max_price = max_price_entry.get()

    if min_price and max_price:
        min_price = int(min_price)
        max_price = int(max_price)
    else:
        min_price = None
        max_price = None

    if query:
        info_text.delete(1.0, tk.END)
        params = {'query': query, 'page_size': 5000, 'page_end': 10, 'min_price': min_price, 'max_price': max_price}
        folder_path = asyncio.run(add_data_to_excel(None, params))
        update_info(f"Данные сохранены в файле: {folder_path}")
    else:
        update_info("Введите запрос!")

def clear_fields():
    query_combobox.set('')
    min_price_entry.delete(0, tk.END)
    max_price_entry.delete(0, tk.END)

def update_info(message):
    info_text.configure(state='normal')
    info_text.insert(tk.END, message + "\n")
    info_text.configure(state='disabled')

def open_folder():
    query = query_combobox.get()
    min_price = int(min_price_entry.get() or 0)
    max_price = int(max_price_entry.get() or 0)
    params = {'query': query, 'page_size': 10, 'page_end': 10, 'min_price': min_price, 'max_price': max_price}
    asyncio.run(add_data_to_excel(None, params))
    folder_path = asyncio.run(add_data_to_excel(None, params))
    subprocess.Popen(f'explorer.exe "{folder_path}"')

# Создание главного окна
root = tk.Tk()
root.title("Экспорт данных с Plati.io")
root.geometry("800x600")
root.resizable(False, False)

# Создание стилей для элементов
style = ttk.Style()
style.configure('TCombobox', fieldbackground='white', font=('Helvetica', 12), foreground='black')
style.configure('TEntry', fieldbackground='white', font=('Helvetica', 12), foreground='black')
style.configure('TLabel', font=('Helvetica', 12))

# Создание элементов интерфейса
query_label = ttk.Label(root, text="Введите запрос:")
query_label.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="w")

query_combobox = ttk.Combobox(root, width=72, style='TCombobox')
query_combobox.grid(row=0, column=1, padx=10, pady=(10, 5), sticky="ew")

price_frame = ttk.LabelFrame(root, text="Цена")
price_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=(5, 10), sticky="w")

min_price_label = ttk.Label(price_frame, text="Минимальная цена:")
min_price_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

min_price_entry = ttk.Entry(price_frame, width=15, style='TEntry')
min_price_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

max_price_label = ttk.Label(price_frame, text="Максимальная цена:")
max_price_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

max_price_entry = ttk.Entry(price_frame, width=15, style='TEntry')
max_price_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

run_button = ttk.Button(root, text="Запустить", command=run_query)
run_button.grid(row=2, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="ew")

clear_button = ttk.Button(root, text="Очистить", command=clear_fields)
clear_button.grid(row=3, column=0, columnspan=2, padx=10, pady=(5, 10), sticky="ew")

open_folder_button = ttk.Button(root, text="Открыть папку с файлами Excel", command=open_folder)
open_folder_button.grid(row=4, column=0, columnspan=2, padx=10, pady=(5, 10), sticky="ew")

info_text = tk.Text(root, wrap="word", height=10, width=72, state='disabled')
info_text.grid(row=5, column=0, columnspan=2, padx=10, pady=(5, 10), sticky="nsew")

# Запуск главного цикла
root.mainloop()